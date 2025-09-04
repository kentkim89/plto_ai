import streamlit as st
import pandas as pd
import io
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from datetime import datetime
import json
import traceback
import requests
from io import BytesIO
import base64

# --------------------------------------------------------------------------
# 페이지 설정 (가장 먼저 실행)
# --------------------------------------------------------------------------
st.set_page_config(
    page_title="주문 처리 자동화 Pro v2.3",
    layout="wide",
    page_icon="📊",
    initial_sidebar_state="expanded"
)

# --------------------------------------------------------------------------
# 라이브러리 가용성 체크
# --------------------------------------------------------------------------

# Microsoft Graph API 사용
GRAPH_AVAILABLE = False
try:
    import msal
    GRAPH_AVAILABLE = True
except ImportError:
    pass

# Plotly 사용
PLOTLY_AVAILABLE = False
try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    pass

# Gemini AI 사용
GEMINI_AVAILABLE = False
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    pass

# --------------------------------------------------------------------------
# Microsoft Graph API 연결 함수
# --------------------------------------------------------------------------

@st.cache_resource
def get_graph_token():
    """Microsoft Graph API 토큰 획득"""
    if not GRAPH_AVAILABLE:
        return None
    
    try:
        if "sharepoint" not in st.secrets:
            return None
        
        tenant_id = st.secrets["sharepoint"]["tenant_id"]
        client_id = st.secrets["sharepoint"]["client_id"]
        client_secret = st.secrets["sharepoint"]["client_secret"]
        
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )
        
        result = app.acquire_token_silent(["https://graph.microsoft.com/.default"], account=None)
        if not result:
            result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        
        if "access_token" in result:
            return result["access_token"]
        else:
            st.error(f"토큰 획득 실패: {result.get('error_description', 'Unknown error')}")
            return None
            
    except Exception as e:
        st.error(f"Graph API 연결 실패: {e}")
        return None

@st.cache_data(ttl=600)
def load_master_data_from_sharepoint():
    """Microsoft Graph API를 통해 SharePoint에서 마스터 데이터 로드"""
    if not GRAPH_AVAILABLE:
        return pd.DataFrame()

    try:
        token = get_graph_token()
        if not token:
            return pd.DataFrame()
        
        headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
        site_url = "https://graph.microsoft.com/v1.0/sites/goremi.sharepoint.com:/sites/data"
        site_response = requests.get(site_url, headers=headers)
        
        if site_response.status_code == 200:
            site_id = site_response.json()['id']
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            drives_response = requests.get(drives_url, headers=headers)
            
            if drives_response.status_code == 200:
                drives = drives_response.json()['value']
                for drive in drives:
                    drive_id = drive['id']
                    search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='plto_master_data.xlsx')"
                    search_response = requests.get(search_url, headers=headers)
                    
                    if search_response.status_code == 200:
                        items = search_response.json().get('value', [])
                        for item in items:
                            if item['name'] == 'plto_master_data.xlsx':
                                download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item['id']}/content"
                                file_response = requests.get(download_url, headers=headers)
                                if file_response.status_code == 200:
                                    st.session_state['sharepoint_site_id'] = site_id
                                    st.session_state['sharepoint_drive_id'] = drive_id
                                    df_master = pd.read_excel(io.BytesIO(file_response.content))
                                    df_master = df_master.drop_duplicates(subset=['SKU코드'], keep='first')
                                    st.success("✅ Microsoft Graph API로 마스터 데이터 로드 성공!")
                                    return df_master
        st.error("Graph API를 통해 마스터 데이터를 찾는 데 실패했습니다.")
    except Exception as e:
        st.error(f"❌ 마스터 데이터 로드 실패: {e}")
    
    return pd.DataFrame()

def save_to_sharepoint_records(df_main_result, df_ecount_upload):
    """Microsoft Graph API를 통해 처리 결과를 SharePoint에 기록/누적합니다."""
    if not GRAPH_AVAILABLE:
        st.info("Graph API가 활성화되지 않아 SharePoint에 자동 저장할 수 없습니다.")
        return False, "Graph API 비활성화"
    
    token = get_graph_token()
    if not token:
        return False, "SharePoint 인증 토큰 획득 실패"
        
    if "sharepoint_drive_id" not in st.session_state:
        st.warning("SharePoint 드라이브 정보를 찾을 수 없습니다. 마스터 데이터를 먼저 로드해야 합니다.")
        load_master_data_from_sharepoint()
        if "sharepoint_drive_id" not in st.session_state:
            return False, "SharePoint 드라이브 정보 로드 실패"

    drive_id = st.session_state['sharepoint_drive_id']
    file_name = "plto_record_data.xlsx"
    file_path_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_name}"

    headers = {'Authorization': f'Bearer {token}'}
    
    existing_df = pd.DataFrame()
    try:
        download_response = requests.get(f"{file_path_url}:/content", headers=headers)
        
        if download_response.status_code == 200:
            if download_response.content and len(download_response.content) > 0:
                try:
                    existing_df = pd.read_excel(io.BytesIO(download_response.content))
                    st.info(f"기존 레코드 '{file_name}'에서 {len(existing_df)}개 데이터를 불러왔습니다.")
                except Exception as e:
                    st.warning(f"'{file_name}' 파일을 읽는 중 오류 발생: {e}. 새 데이터로 덮어씁니다.")
            else:
                st.info(f"기존 레코드 '{file_name}' 파일이 비어 있어 새로 데이터를 입력합니다.")
        elif download_response.status_code == 404:
            st.info(f"기존 레코드 '{file_name}'을 찾을 수 없어 새로 생성합니다.")
        else:
            error_details = download_response.json()
            return False, f"기존 레코드 다운로드 실패 ({download_response.status_code}): {error_details.get('error', {}).get('message', '알 수 없는 오류')}"
    except Exception as e:
        return False, f"기존 레코드 처리 중 오류: {e}"

    try:
        order_date_str = None
        if not df_ecount_upload.empty:
            first_date_val = df_ecount_upload['일자'].iloc[0]
            if pd.notna(first_date_val) and str(first_date_val).strip():
                order_date_str = str(first_date_val)

        if not order_date_str:
            order_date_str = datetime.now().strftime("%Y%m%d")
            st.info("이카운트 데이터에서 유효한 '일자'를 찾을 수 없어 오늘 날짜를 사용합니다.")
        
        try:
            order_date = pd.to_datetime(order_date_str, format='%Y%m%d').strftime('%Y-%m-%d')
        except ValueError:
            st.warning(f"날짜 형식('{order_date_str}')이 올바르지 않아 오늘 날짜로 대체합니다.")
            order_date = datetime.now().strftime('%Y-%m-%d')

        new_records = pd.DataFrame({
            '주문일자': order_date,
            '처리일시': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            '재고관리코드': df_main_result['재고관리코드'],
            'SKU상품명': df_main_result['SKU상품명'],
            '주문수량': df_main_result['주문수량'],
            '실결제금액': df_main_result['실결제금액'],
            '쇼핑몰': df_main_result['쇼핑몰'],
            '수령자명': df_main_result['수령자명']
        })
    
        combined_df = pd.concat([existing_df, new_records], ignore_index=True)
    
        output = BytesIO()
        combined_df.to_excel(output, index=False, sheet_name='Records')
        file_content = output.getvalue()
    
        upload_headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        upload_response = requests.put(f"{file_path_url}:/content", headers=upload_headers, data=file_content)
        
        if upload_response.status_code in [200, 201]:
            return True, f"✅ SharePoint에 {len(new_records)}개 신규 레코드 저장 완료 (총 {len(combined_df)}개)"
        else:
            error_details = upload_response.json()
            return False, f"SharePoint 업로드 실패 ({upload_response.status_code}): {error_details.get('error', {}).get('message', '알 수 없는 오류')}"
    except Exception as e:
        st.error(traceback.format_exc())
        return False, f"레코드 저장 중 오류 발생: {e}"

# --------------------------------------------------------------------------
# AI 분석 함수
# --------------------------------------------------------------------------

@st.cache_resource
def init_gemini():
    """Gemini AI 초기화. 여러 모델을 시도하여 안정성을 높입니다."""
    if not GEMINI_AVAILABLE:
        return None
    
    try:
        api_key = st.secrets.get("GEMINI_API_KEY")
        if not api_key:
            st.warning("Gemini API 키를 찾을 수 없습니다.")
            return None
            
        genai.configure(api_key=api_key)
        
        model_candidates = ['gemini-1.5-flash-latest', 'gemini-1.0-pro', 'gemini-pro']
        for model_name in model_candidates:
            try:
                model = genai.GenerativeModel(model_name)
                st.session_state['gemini_model_name'] = model_name
                return model
            except Exception:
                st.info(f"'{model_name}' 모델 초기화 실패. 다음 모델을 시도합니다.")
        
        st.error("사용 가능한 Gemini 모델을 찾지 못했습니다.")
        return None
    except Exception as e:
        st.error(f"Gemini AI 초기화 중 오류 발생: {e}")
        return None

def analyze_sales_with_ai(df_records):
    """AI를 사용한 판매 데이터 분석"""
    if not GEMINI_AVAILABLE:
        return "AI 분석이 비활성화되어 있습니다."
    
    model = init_gemini()
    if not model or df_records.empty:
        return "AI 모델을 초기화할 수 없거나 분석할 데이터가 없습니다."
        
    try:
        summary = {
            "총_주문수": len(df_records),
            "총_매출": float(df_records['실결제금액'].sum()),
            "상품_종류": int(df_records['SKU상품명'].nunique()),
            "고객수": int(df_records['수령자명'].nunique()),
            "베스트셀러_TOP5": df_records.groupby('SKU상품명')['주문수량'].sum().nlargest(5).to_dict(),
            "채널별_매출": {k: float(v) for k, v in df_records.groupby('쇼핑몰')['실결제금액'].sum().to_dict().items()}
        }
        
        prompt = f"""
        온라인 쇼핑몰 판매 데이터를 분석해주세요: {json.dumps(summary, ensure_ascii=False, indent=2, default=str)}
        다음 내용을 포함하여 분석해주세요:
        1. 📈 판매 트렌드 분석
        2. 🏆 베스트셀러 인사이트
        3. 🛒 채널별 성과 평가
        4. 💡 실행 가능한 개선 제안
        분석은 구체적이고 실용적으로 작성해주세요.
        """
        
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        model_name = st.session_state.get('gemini_model_name', '알 수 없음')
        return f"AI 분석 오류 ({model_name} 모델 사용 중): {e}"

# --------------------------------------------------------------------------
# 데이터 처리 함수
# --------------------------------------------------------------------------

def to_excel_formatted(df, format_type=None):
    """데이터프레임을 서식이 적용된 엑셀 파일로 변환"""
    output = BytesIO()
    df_to_save = df.copy().fillna('')
    
    if format_type == 'ecount_upload':
        df_to_save = df_to_save.rename(columns={'적요_전표': '적요', '적요_품목': '적요.1'})

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_to_save.to_excel(writer, index=False, sheet_name='Sheet1')
    
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active

    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    for column_cells in sheet.columns:
        try:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            adjusted_width = min((max_length + 2) * 1.2, 50)
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        except (ValueError, TypeError):
            pass

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    pink_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border

    if format_type == 'packing_list':
        bundle_start_row = 2
        for row_num in range(2, sheet.max_row + 2):
            is_new_bundle = (row_num <= sheet.max_row and sheet.cell(row=row_num, column=1).value) or row_num > sheet.max_row
            if is_new_bundle and row_num > 2:
                bundle_end_row = row_num - 1
                try:
                    bundle_num_str = str(sheet.cell(row=bundle_start_row, column=1).value)
                    if bundle_num_str.isdigit() and int(bundle_num_str) % 2 != 0:
                        for r in range(bundle_start_row, bundle_end_row + 1):
                            for c in range(1, sheet.max_column + 1):
                                sheet.cell(row=r, column=c).fill = pink_fill
                except (ValueError, IndexError):
                    pass
                
                if bundle_start_row < bundle_end_row:
                    sheet.merge_cells(start_row=bundle_start_row, start_column=1, end_row=bundle_end_row, end_column=1)
                    sheet.merge_cells(start_row=bundle_start_row, start_column=4, end_row=bundle_end_row, end_column=4)
                bundle_start_row = row_num
    elif format_type == 'quantity_summary':
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row)):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = pink_fill
    
    final_output = BytesIO()
    workbook.save(final_output)
    return final_output.getvalue()

def process_all_files(file1, file2, file3, df_master):
    """메인 처리 함수"""
    try:
        df_smartstore = pd.read_excel(file1)
        df_ecount_orig = pd.read_excel(file2)
        df_godomall = pd.read_excel(file3)

        df_ecount_orig['original_order'] = range(len(df_ecount_orig))
        
        df_godomall.rename(columns={'회 할인 금액': '회원 할인 금액', '자체옵션코드': '재고관리코드'}, inplace=True, errors='ignore')
        
        cols_to_numeric = ['상품별 품목금액', '총 배송 금액', '회원 할인 금액', '쿠폰 할인 금액', '사용된 마일리지', '총 결제 금액']
        for col in cols_to_numeric:
            if col in df_godomall.columns:
                df_godomall[col] = pd.to_numeric(df_godomall[col].astype(str).str.replace('[원,]', '', regex=True), errors='coerce').fillna(0)
        
        df_godomall['보정된_배송비'] = np.where(df_godomall.duplicated(subset=['수취인 이름']), 0, df_godomall.get('총 배송 금액', 0))
        df_godomall['수정될_금액_고도몰'] = (df_godomall.get('상품별 품목금액', 0) + df_godomall['보정된_배송비'] - 
                                     df_godomall.get('회원 할인 금액', 0) - df_godomall.get('쿠폰 할인 금액', 0) - 
                                     df_godomall.get('사용된 마일리지', 0))
        
        warnings = [f"- [금액 불일치] **{name}**님: {group['수정될_금액_고도몰'].sum() - group['총 결제 금액'].iloc[0]:,.0f}원 차이" for name, group in df_godomall.groupby('수취인 이름') if '총 결제 금액' in group and abs(group['수정될_금액_고도몰'].sum() - group['총 결제 금액'].iloc[0]) > 1]

        df_final = df_ecount_orig.copy().rename(columns={'금액': '실결제금액'})
        
        key_cols = ['재고관리코드', '주문수량', '수령자명']
        smartstore_prices = df_smartstore.rename(columns={'실결제금액': '수정될_금액_스토어'})[key_cols + ['수정될_금액_스토어']].drop_duplicates(subset=key_cols)
        godomall_prices = df_godomall.rename(columns={'수취인 이름': '수령자명', '상품수량': '주문수량'})[key_cols + ['수정될_금액_고도몰']].drop_duplicates(subset=key_cols)
        
        for df in [df_final, smartstore_prices, godomall_prices]:
            for col in ['재고관리코드', '수령자명']: df[col] = df[col].astype(str).str.strip()
            for col in ['주문수량']: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        
        df_final = pd.merge(df_final, smartstore_prices, on=key_cols, how='left')
        df_final = pd.merge(df_final, godomall_prices, on=key_cols, how='left')

        df_final['실결제금액'] = np.where(df_final['쇼핑몰'] == '고도몰5', df_final['수정될_금액_고도몰'], df_final['실결제금액'])
        df_final['실결제금액'] = np.where(df_final['쇼핑몰'] == '스마트스토어', df_final['수정될_금액_스토어'], df_final['실결제금액'])
        df_final['실결제금액'].fillna(df_ecount_orig['금액'], inplace=True)
        
        df_main_result = df_final[['재고관리코드', 'SKU상품명', '주문수량', '실결제금액', '쇼핑몰', '수령자명', 'original_order']]
        
        name_groups = df_main_result.groupby('수령자명')['original_order'].apply(list)
        warnings.extend([f"- [동명이인 의심] **{name}**님의 주문이 떨어져 있습니다." for name, orders in name_groups.items() if len(orders) > 1 and (max(orders) - min(orders) + 1) != len(orders)])
        
        df_quantity_summary = df_main_result.groupby('SKU상품명', as_index=False)['주문수량'].sum().rename(columns={'주문수량': '개수'})
        
        df_packing = df_main_result.sort_values('original_order')[['SKU상품명', '주문수량', '수령자명', '쇼핑몰']].copy()
        is_first = ~df_packing['수령자명'].duplicated(keep='first')
        df_packing['묶음번호'] = is_first.cumsum()
        df_packing_list = df_packing.copy()
        df_packing_list['묶음번호'] = df_packing_list['묶음번호'].where(is_first, '')
        df_packing_list = df_packing_list[['묶음번호', 'SKU상품명', '주문수량', '수령자명', '쇼핑몰']]

        df_merged = pd.merge(df_main_result, df_master[['SKU코드', '과세여부', '입수량']], left_on='재고관리코드', right_on='SKU코드', how='left')
        warnings.extend([f"- [미등록] {row['재고관리코드']}: {row['SKU상품명']}" for _, row in df_merged[df_merged['SKU코드'].isna()].iterrows()])

        client_map = {'쿠팡': '쿠팡 주식회사', '고도몰5': '고래미자사몰_현금영수증(고도몰)', '스마트스토어': '스토어팜', '배민상회': '주식회사 우아한형제들(배민상회)', '이지웰몰': '주식회사 현대이지웰'}
        
        df_ecount = pd.DataFrame()
        df_ecount['일자'] = datetime.now().strftime("%Y%m%d")
        df_ecount['거래처명'] = df_merged['쇼핑몰'].map(client_map).fillna(df_merged['쇼핑몰'])
        df_ecount['출하창고'] = '고래미'
        df_ecount['거래유형'] = np.where(df_merged['과세여부'] == '면세', 12, 11)
        df_ecount['적요_전표'] = '오전/온라인'
        df_ecount['품목코드'] = df_merged['재고관리코드']
        
        is_box = df_merged['SKU상품명'].str.contains("BOX", na=False)
        입수량 = pd.to_numeric(df_merged['입수량'], errors='coerce').fillna(1)
        base_qty = np.where(is_box, df_merged['주문수량'] * 입수량, df_merged['주문수량'])
        is_3pack = df_merged['SKU상품명'].str.contains("3개입|3개", na=False)
        df_ecount['수량'] = np.where(is_3pack, base_qty * 3, base_qty).astype(int)
        df_ecount['박스'] = np.where(is_box, df_merged['주문수량'], np.nan)
        
        df_merged['실결제금액'] = pd.to_numeric(df_merged['실결제금액'], errors='coerce').fillna(0)
        공급가액 = np.where(df_merged['과세여부'] == '과세', df_merged['실결제금액'] / 1.1, df_merged['실결제금액'])
        
        df_ecount['공급가액'] = 공급가액.round().astype(int)
        df_ecount['부가세'] = (df_merged['실결제금액'] - df_ecount['공급가액']).round().astype(int)
        
        df_ecount['쇼핑몰고객명'] = df_merged['수령자명']
        df_ecount['original_order'] = df_merged['original_order']
        
        sort_order = ['고래미자사몰_현금영수증(고도몰)', '스토어팜', '쿠팡 주식회사', '주식회사 우아한형제들(배민상회)', '주식회사 현대이지웰']
        df_ecount['거래처명_sort'] = pd.Categorical(df_ecount['거래처명'], categories=sort_order, ordered=True)
        df_ecount = df_ecount.sort_values(by=['거래처명_sort', '거래유형', 'original_order']).drop(columns=['거래처명_sort', 'original_order'])
        
        ecount_columns = ['일자', '순번', '거래처코드', '거래처명', '담당자', '출하창고', '거래유형', '통화', '환율', '적요_전표', '미수금', '총합계', '연결전표', '품목코드', '품목명', '규격', '박스', '수량', '단가', '외화금액', '공급가액', '부가세', '적요_품목', '생산전표생성', '시리얼/로트', '관리항목', '쇼핑몰고객명']
        df_ecount_upload = df_ecount.reindex(columns=ecount_columns, fill_value='')

        return (df_main_result.drop(columns=['original_order']), df_quantity_summary, df_packing_list, df_ecount_upload, True, "✅ 모든 처리가 완료되었습니다!", warnings)

    except Exception as e:
        st.error(traceback.format_exc())
        return None, None, None, None, False, f"❌ 오류: {str(e)}", []

def create_analytics_dashboard(df_records):
    """분석 대시보드 생성"""
    if df_records.empty:
        st.warning("분석할 데이터가 없습니다.")
        return
    
    st.header("판매 데이터 분석")
    col1, col2, col3, col4 = st.columns(4)
    total_revenue = df_records['실결제금액'].sum()
    total_orders = len(df_records['수령자명'].unique())
    col1.metric("💰 총 매출", f"₩{total_revenue:,.0f}")
    col2.metric("📦 총 주문수", f"{total_orders:,}")
    col3.metric("💵 평균 주문액", f"₩{total_revenue/total_orders if total_orders else 0:,.0f}")
    col4.metric("👥 고객수", f"{df_records['수령자명'].nunique():,}")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📈 일별 트렌드", "🏆 베스트셀러", "🛒 채널 분석", "🤖 AI 인사이트"])
    
    with tab1:
        st.subheader("일별 매출 트렌드")
        df_records['주문일자'] = pd.to_datetime(df_records['주문일자'])
        daily_sales = df_records.groupby(df_records['주문일자'].dt.date)['실결제금액'].sum()
        st.line_chart(daily_sales)
    
    with tab2:
        st.subheader("상품별 판매 TOP 10")
        st.bar_chart(df_records.groupby('SKU상품명')['주문수량'].sum().nlargest(10))
    
    with tab3:
        st.subheader("채널별 매출")
        st.bar_chart(df_records.groupby('쇼핑몰')['실결제금액'].sum())
    
    with tab4:
        if GEMINI_AVAILABLE:
            with st.spinner("🤖 AI가 데이터를 분석 중입니다..."):
                st.markdown("### 🤖 AI 판매 분석 리포트")
                st.markdown(analyze_sales_with_ai(df_records))
        else:
            st.warning("AI 분석 기능을 사용하려면 google-generativeai를 설치하세요.")

# --------------------------------------------------------------------------
# 메인 앱
# --------------------------------------------------------------------------
def main():
    with st.sidebar:
        st.title("📊 Order Pro v2.3")
        st.markdown("---")
        menu = st.radio("메뉴 선택", ["📑 주문 처리", "📈 판매 분석", "⚙️ 설정"])
        st.markdown("---")
        st.caption("📌 시스템 상태")
        st.success("✅ Graph API 연결" if GRAPH_AVAILABLE and get_graph_token() else ("⚠️ Graph API 인증 필요" if GRAPH_AVAILABLE else "💾 로컬 모드"))
        st.success("✅ AI 활성화" if GEMINI_AVAILABLE else "🤖 AI 비활성화")
        if st.button("🔄 캐시 초기화"):
            st.cache_data.clear()
            st.cache_resource.clear()
            st.success("캐시 초기화 완료")
            st.rerun()

    if menu == "📑 주문 처리":
        st.title("📑 주문 처리 자동화")
        st.info("💡 SharePoint 연동 및 처리 결과 자동 누적 기능이 활성화되어 있습니다.")
        
        with st.expander("📊 마스터 데이터 상태", expanded=True):
            df_master = load_master_data_from_sharepoint()
            if df_master.empty:
                st.warning("⚠️ SharePoint에서 마스터 데이터를 불러오지 못했습니다. 로컬 업로드를 사용하세요.")
                uploaded_master = st.file_uploader("마스터 데이터 업로드 (xlsx, xls, csv)", type=['xlsx', 'xls', 'csv'])
                if uploaded_master:
                    try:
                        df_master = pd.read_excel(uploaded_master) if uploaded_master.name.endswith(('xlsx', 'xls')) else pd.read_csv(uploaded_master)
                        df_master = df_master.drop_duplicates(subset=['SKU코드'], keep='first')
                        st.success(f"✅ 로컬 마스터 데이터 {len(df_master)}개 로드 완료")
                    except Exception as e:
                        st.error(f"파일을 읽는 데 실패했습니다: {e}")
                        df_master = pd.DataFrame()
            
            if not df_master.empty:
                col1, col2, col3 = st.columns(3)
                col1.metric("총 SKU", f"{len(df_master):,}개")
                col2.metric("과세 상품", f"{(df_master['과세여부']=='과세').sum():,}개")
                col3.metric("면세 상품", f"{(df_master['과세여부']=='면세').sum():,}개")

        st.markdown("---")
        st.header("1️⃣ 원본 파일 업로드")
        col1, col2, col3 = st.columns(3)
        file1 = col1.file_uploader("스마트스토어", type=['xlsx', 'xls'])
        file2 = col2.file_uploader("이카운트", type=['xlsx', 'xls'])
        file3 = col3.file_uploader("고도몰", type=['xlsx', 'xls'])
        
        st.header("2️⃣ 데이터 처리")
        if st.button("🚀 처리 시작", type="primary", disabled=not all([file1, file2, file3, not df_master.empty])):
            with st.spinner('처리 중...'):
                result = process_all_files(file1, file2, file3, df_master)
                df_main, df_qty, df_pack, df_ecount, success, message, warnings = result
            
            if success:
                st.balloons()
                st.success(message)
                
                if GRAPH_AVAILABLE:
                    with st.spinner('SharePoint에 기록 저장 중...'):
                        save_success, save_msg = save_to_sharepoint_records(df_main, df_ecount)
                        st.success(save_msg) if save_success else st.warning(save_msg)
                
                if warnings:
                    with st.expander(f"⚠️ 확인 필요 ({len(warnings)}건)"):
                        st.markdown("\n".join(warnings))
                
                st.session_state['last_result'] = df_main
                st.session_state['processed_date'] = df_ecount['일자'].iloc[0]
                
                st.markdown("---")
                st.header("3️⃣ 결과 다운로드")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                tabs = st.tabs(["🏢 이카운트", "📋 포장리스트", "📦 수량요약", "✅ 최종결과"])
                
                tabs[0].dataframe(df_ecount.head(20), use_container_width=True)
                tabs[0].download_button("📥 이카운트 업로드용 다운로드", to_excel_formatted(df_ecount, 'ecount_upload'), f"이카운트_{timestamp}.xlsx")
                
                tabs[1].dataframe(df_pack.head(20), use_container_width=True)
                tabs[1].download_button("📥 포장리스트 다운로드", to_excel_formatted(df_pack, 'packing_list'), f"포장리스트_{timestamp}.xlsx")
                
                tabs[2].dataframe(df_qty, use_container_width=True)
                tabs[2].download_button("📥 수량요약 다운로드", to_excel_formatted(df_qty, 'quantity_summary'), f"수량요약_{timestamp}.xlsx")

                tabs[3].dataframe(df_main.head(20), use_container_width=True)
                tabs[3].download_button("📥 최종결과 다운로드", to_excel_formatted(df_main), f"최종결과_{timestamp}.xlsx")
            else:
                st.error(message)
    
    elif menu == "📈 판매 분석":
        st.title("📈 판매 분석")
        if 'last_result' in st.session_state and not st.session_state['last_result'].empty:
            df_records = st.session_state['last_result'].copy()
            df_records['주문일자'] = pd.to_datetime(st.session_state['processed_date'], format='%Y%m%d')
            create_analytics_dashboard(df_records)
        else:
            st.info("먼저 '주문 처리' 메뉴에서 데이터를 처리해주세요.")
    
    elif menu == "⚙️ 설정":
        st.title("⚙️ 시스템 설정")
        st.header("📁 Microsoft Graph API 설정")
        if "sharepoint" in st.secrets:
            st.text_input("Tenant ID", value=st.secrets["sharepoint"].get("tenant_id", "")[:20]+"...", disabled=True)
            st.text_input("Client ID", value=st.secrets["sharepoint"].get("client_id", "")[:20]+"...", disabled=True)
            if st.button("🔄 Graph API 연결 테스트"):
                with st.spinner("테스트 중..."):
                    st.success("✅ Microsoft Graph API 연결 성공!") if get_graph_token() else st.error("❌ Graph API 연결 실패")
        else:
            st.warning("Graph API 설정이 없습니다. (st.secrets.sharepoint)")
        
        st.header("🤖 AI 설정")
        if "GEMINI_API_KEY" in st.secrets:
            st.text_input("Gemini API Key", value=st.secrets["GEMINI_API_KEY"][:10]+"...", disabled=True)
            if st.button("🔄 AI 연결 테스트"):
                with st.spinner("테스트 중..."):
                    model = init_gemini()
                    st.success(f"✅ Gemini AI 연결 성공! (모델: {st.session_state.get('gemini_model_name', 'N/A')})") if model else st.error("❌ AI 연결 실패")
        else:
            st.warning("Gemini API 키가 설정되지 않았습니다. (st.secrets.GEMINI_API_KEY)")

if __name__ == "__main__":
    main()
